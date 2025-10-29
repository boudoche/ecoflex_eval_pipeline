import os
import json
import logging
from typing import Any, Dict, List, Optional, Tuple
import smtplib
import ssl
from email.message import EmailMessage

from fastapi import FastAPI, HTTPException, Query, Header, Request
from fastapi.middleware.cors import CORSMiddleware
from fastapi.staticfiles import StaticFiles
from fastapi.responses import RedirectResponse

from evaluate import (
    load_questions,
    evaluate_submission,
    write_results,
)
from reporting import write_summary_csv

import asyncio
from concurrent.futures import ThreadPoolExecutor

QUESTIONS_PATH = os.getenv("QUESTIONS_PATH", os.path.join(os.path.dirname(__file__), "questions.json"))
RESULTS_DIR = os.getenv("RESULTS_DIR", os.path.join(os.path.dirname(__file__), "results"))
# Fixed OpenAI model
DEFAULT_MODEL = "gpt-4o-mini"
# Fixed number of parallel workers for grading
FIXED_WORKERS = int(os.getenv("FIXED_WORKERS", "6"))
# Token sources
TOKENS_PATH = os.getenv("TOKENS_PATH", os.path.join(os.path.dirname(__file__), "tokens.json"))
TEAM_TOKENS = os.getenv("TEAM_TOKENS", "")  # format: token:Team[:email],token:Team[:email]
# Max submission size (in bytes, default 5MB)
MAX_SUBMISSION_SIZE = int(os.getenv("MAX_SUBMISSION_SIZE", str(5 * 1024 * 1024)))

# Logging configuration (includes filename and line number)
LOG_LEVEL = os.getenv("LOG_LEVEL", "INFO").upper()
logging.basicConfig(
    level=LOG_LEVEL,
    format="%(asctime)s %(levelname)s %(name)s [%(process)d] %(filename)s:%(lineno)d - %(message)s",
    datefmt="%Y-%m-%dT%H:%M:%S.%f%z",
)
logger = logging.getLogger("ecoflex")

app = FastAPI(title="Ecoflex Auto Grader", version="1.3.3")

# Allow CORS for simple integration/testing; tighten in production
app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"]
)

# Mount static UI
ui_dir = os.path.join(os.path.dirname(__file__), "ui")
if os.path.isdir(ui_dir):
    app.mount("/ui", StaticFiles(directory=ui_dir, html=True), name="ui")

@app.get("/")
async def root_redirect() -> RedirectResponse:
    return RedirectResponse(url="/ui/")

# In-memory state
_state: Dict[str, Any] = {
    "questions": {},
    "token_to_info": {},  # token -> {team:str, email:str, used:bool}
    "executor": None,
}


def _ensure_results_dir() -> None:
    os.makedirs(RESULTS_DIR, exist_ok=True)


def _load_tokens() -> Dict[str, Dict[str, Any]]:
    result: Dict[str, Dict[str, Any]] = {}
    # From env: token:Team[:email]
    env_value = os.getenv("TEAM_TOKENS", TEAM_TOKENS)
    if env_value:
        parts = [p.strip() for p in env_value.split(",") if p.strip()]
        for part in parts:
            fields = part.split(":")
            if len(fields) >= 2:
                token = fields[0].strip()
                team = fields[1].strip()
                email = fields[2].strip() if len(fields) >= 3 else ""
                if token:
                    result[token] = {"team": team, "email": email, "used": False}
    # From file: either {token: team} or {token: {team, email, used}}
    path_value = os.getenv("TOKENS_PATH", TOKENS_PATH)
    if os.path.isfile(path_value):
        try:
            with open(path_value, "r", encoding="utf-8") as f:
                data = json.load(f)
            if isinstance(data, dict):
                for token, val in data.items():
                    if not isinstance(token, str):
                        continue
                    if isinstance(val, str):
                        result[token] = {"team": val, "email": result.get(token, {}).get("email", ""), "used": False}
                    elif isinstance(val, dict):
                        team = str(val.get("team", result.get(token, {}).get("team", "")))
                        email = str(val.get("email", result.get(token, {}).get("email", "")))
                        used = bool(val.get("used", False))
                        if team:
                            result[token] = {"team": team, "email": email, "used": used}
        except Exception:
            pass
    return result


def _persist_tokens(mapping: Dict[str, Dict[str, Any]]) -> None:
    path_value = os.getenv("TOKENS_PATH", TOKENS_PATH)
    if not path_value:
        return
    # Only persist if a path is provided; write full mapping
    try:
        os.makedirs(os.path.dirname(path_value) or ".", exist_ok=True)
        with open(path_value, "w", encoding="utf-8") as f:
            json.dump(mapping, f, indent=2, ensure_ascii=False)
            f.write("\n")
    except Exception:
        logger.exception("Failed to persist tokens file at %s", path_value)


def _require_token_and_team(x_submission_token: Optional[str]) -> Tuple[str, Dict[str, Any]]:
    if not x_submission_token:
        raise HTTPException(status_code=401, detail="Missing submission token")
    info = _state["token_to_info"].get(x_submission_token)
    if not info or not info.get("team"):
        raise HTTPException(status_code=401, detail="Invalid submission token")
    # Enforce one submission per token
    if bool(info.get("used", False)):
        raise HTTPException(status_code=409, detail="Submission already received for this token")
    return x_submission_token, info


def _send_confirmation_email(to_addr: str, participant_id: str, submission_json: Dict[str, Any]) -> None:
    if not to_addr:
        return
    if os.getenv("EMAIL_ENABLED", "").lower() not in ("1", "true", "yes", "on"):
        return
    host = os.getenv("SMTP_HOST", "")
    port = int(os.getenv("SMTP_PORT", "587"))
    user = os.getenv("SMTP_USER", "")
    password = os.getenv("SMTP_PASS", "")
    from_addr = os.getenv("SMTP_FROM", user)
    from_name = os.getenv("SMTP_FROM_NAME", "Argusa Data Challenge")
    reply_to = os.getenv("SMTP_REPLY_TO", from_addr)
    use_ssl = os.getenv("SMTP_USE_SSL", "").lower() in ("1", "true", "yes", "on")
    if not host or not from_addr:
        logger.warning("Email disabled: SMTP_HOST/SMTP_FROM not configured")
        return
    
    # Build email with proper headers to avoid spam
    msg = EmailMessage()
    msg["From"] = f"{from_name} <{from_addr}>"
    msg["To"] = to_addr
    msg["Reply-To"] = reply_to
    msg["Subject"] = f"✓ Submission Received - {participant_id}"
    
    # Add additional headers to improve deliverability
    msg["X-Mailer"] = "Ecoflex Auto Grader"
    msg["X-Priority"] = "3"  # Normal priority
    msg["Importance"] = "Normal"
    
    # Create a more professional email body (plain text + HTML)
    text_body = f"""Hello {participant_id},

Your submission has been received successfully!

Submission Details:
- Team/Participant: {participant_id}
- Number of answers: {len(submission_json.get('answers', []))}
- Status: Received

Your submission file is attached to this email for your records.

We will notify you once the evaluation is complete.

Best regards,
The Argusa Data Challenge Team

---
This is an automated message. Please do not reply to this email.
"""
    
    # Get logo URL from environment (optional)
    logo_url = os.getenv("EMAIL_LOGO_URL", "")
    
    html_body = f"""<!DOCTYPE html>
<html>
<head>
    <meta charset="utf-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
</head>
<body style="margin: 0; padding: 0; font-family: Arial, sans-serif; background-color: #f4f4f4;">
    <!-- Main container table -->
    <table role="presentation" cellspacing="0" cellpadding="0" border="0" width="100%" style="background-color: #f4f4f4;">
        <tr>
            <td style="padding: 20px 0;">
                <!-- Content wrapper -->
                <table role="presentation" cellspacing="0" cellpadding="0" border="0" width="600" style="margin: 0 auto; background-color: #ffffff;" align="center">
                    
                    <!-- Header with Argusa brand color -->
                    <tr>
                        <td style="background-color: #004B87; padding: 30px 20px; text-align: center; color: #ffffff;">
                            {"<img src='" + logo_url + "' alt='Argusa Logo' style='max-width: 200px; height: auto; margin-bottom: 15px; display: block; margin-left: auto; margin-right: auto;' />" if logo_url else ""}
                            <h1 style="margin: 0; font-size: 24px; font-weight: bold; color: #ffffff;">✓ Submission Received</h1>
                        </td>
                    </tr>
                    
                    <!-- Brand color bar (Argusa colors: Blue, Yellow, Black, Orange) -->
                    <tr>
                        <td style="padding: 0;">
                            <table role="presentation" cellspacing="0" cellpadding="0" border="0" width="100%">
                                <tr>
                                    <td width="25%" style="background-color: #004B87; height: 8px;"></td>
                                    <td width="25%" style="background-color: #FDB913; height: 8px;"></td>
                                    <td width="25%" style="background-color: #000000; height: 8px;"></td>
                                    <td width="25%" style="background-color: #E94E1B; height: 8px;"></td>
                                </tr>
                            </table>
                        </td>
                    </tr>
                    
                    <!-- Content section -->
                    <tr>
                        <td style="padding: 30px 20px; background-color: #f9f9f9; border-left: 1px solid #ddd; border-right: 1px solid #ddd;">
                            <p style="margin: 0 0 15px 0; font-size: 16px; color: #333333; line-height: 1.6;">Hello <strong>{participant_id}</strong>,</p>
                            
                            <p style="margin: 0 0 20px 0; font-size: 18px; font-weight: bold; color: #004B87; line-height: 1.6;">Your submission has been received successfully!</p>
                            
                            <!-- Details box -->
                            <table role="presentation" cellspacing="0" cellpadding="0" border="0" width="100%" style="margin: 20px 0;">
                                <tr>
                                    <td style="background-color: #ffffff; border-left: 5px solid #004B87; padding: 20px; border-radius: 3px;">
                                        <h3 style="margin: 0 0 15px 0; font-size: 18px; color: #004B87;">Submission Details</h3>
                                        <table role="presentation" cellspacing="0" cellpadding="0" border="0" width="100%">
                                            <tr>
                                                <td style="padding: 8px 0; border-bottom: 1px solid #f0f0f0;">
                                                    <p style="margin: 0; font-size: 14px; color: #333333; line-height: 1.6;"><strong>Team/Participant:</strong> {participant_id}</p>
                                                </td>
                                            </tr>
                                            <tr>
                                                <td style="padding: 8px 0; border-bottom: 1px solid #f0f0f0;">
                                                    <p style="margin: 0; font-size: 14px; color: #333333; line-height: 1.6;"><strong>Number of answers:</strong> {len(submission_json.get('answers', []))}</p>
                                                </td>
                                            </tr>
                                            <tr>
                                                <td style="padding: 8px 0;">
                                                    <p style="margin: 0; font-size: 14px; color: #333333; line-height: 1.6;"><strong>Status:</strong> Received</p>
                                                </td>
                                            </tr>
                                        </table>
                                    </td>
                                </tr>
                            </table>
                            
                            <p style="margin: 20px 0 15px 0; font-size: 16px; color: #333333; line-height: 1.6;">Your submission file is attached to this email for your records.</p>
                            
                            <p style="margin: 0 0 20px 0; font-size: 16px; color: #333333; line-height: 1.6;">We will notify you once the evaluation is complete.</p>
                            
                            <p style="margin: 20px 0 0 0; font-size: 16px; color: #333333; line-height: 1.6;">Best regards,<br>
                            <strong>The Argusa Data Challenge Team</strong></p>
                        </td>
                    </tr>
                    
                    <!-- Footer -->
                    <tr>
                        <td style="background-color: #f5f5f5; padding: 20px; text-align: center; border-left: 1px solid #ddd; border-right: 1px solid #ddd; border-bottom: 1px solid #ddd;">
                            <!-- Brand color bar -->
                            <table role="presentation" cellspacing="0" cellpadding="0" border="0" width="100%" style="margin-bottom: 10px;">
                                <tr>
                                    <td width="25%" style="background-color: #004B87; height: 8px;"></td>
                                    <td width="25%" style="background-color: #FDB913; height: 8px;"></td>
                                    <td width="25%" style="background-color: #000000; height: 8px;"></td>
                                    <td width="25%" style="background-color: #E94E1B; height: 8px;"></td>
                                </tr>
                            </table>
                            <p style="margin: 0; font-size: 12px; color: #666666; line-height: 1.6;">This is an automated message. Please do not reply to this email.</p>
                        </td>
                    </tr>
                </table>
            </td>
        </tr>
    </table>
</body>
</html>
"""
    
    # Set both plain text and HTML versions
    msg.set_content(text_body)
    msg.add_alternative(html_body, subtype="html")
    
    # Attach submission JSON
    payload = json.dumps(submission_json, indent=2, ensure_ascii=False).encode("utf-8")
    msg.add_attachment(payload, maintype="application", subtype="json", filename=f"{participant_id}_submission.json")
    try:
        if use_ssl:
            context = ssl.create_default_context()
            with smtplib.SMTP_SSL(host, port, context=context, timeout=30) as s:
                if user and password:
                    s.login(user, password)
                s.send_message(msg)
        else:
            with smtplib.SMTP(host, port, timeout=30) as s:
                s.ehlo()
                try:
                    s.starttls(context=ssl.create_default_context())
                    s.ehlo()
                except Exception:
                    pass
                if user and password:
                    s.login(user, password)
                s.send_message(msg)
        logger.info("Sent confirmation email to %s", to_addr)
    except Exception as exc:
        logger.exception("Failed to send confirmation email to %s: %s", to_addr, exc)


@app.on_event("startup")
async def startup_event() -> None:
    _ensure_results_dir()
    try:
        _state["questions"] = load_questions(QUESTIONS_PATH)
    except Exception as exc:
        logger.exception("Failed to load questions from %s", QUESTIONS_PATH)
        raise RuntimeError(f"Failed to load questions from {QUESTIONS_PATH}: {exc}")
    _state["token_to_info"] = _load_tokens()
    # ThreadPool for running CPU/IO bound grading off the event loop
    max_threads = max(4, FIXED_WORKERS)
    _state["executor"] = ThreadPoolExecutor(max_workers=max_threads)


@app.on_event("shutdown")
async def shutdown_event() -> None:
    executor = _state.get("executor")
    if executor is not None:
        executor.shutdown(wait=False, cancel_futures=True)


@app.get("/health")
async def health() -> Dict[str, str]:
    return {"status": "ok"}


@app.post("/reload-questions")
async def reload_questions() -> Dict[str, str]:
    try:
        _state["questions"] = load_questions(QUESTIONS_PATH)
        return {"status": "reloaded"}
    except Exception as exc:
        logger.exception("Failed to reload questions")
        raise HTTPException(status_code=400, detail=f"Failed to reload questions: {exc}")


@app.post("/reload-tokens")
async def reload_tokens() -> Dict[str, int]:
    mapping = _load_tokens()
    _state["token_to_info"] = mapping
    return {"loaded": len(mapping)}


async def _run_in_executor(func, *args, **kwargs):
    loop = asyncio.get_running_loop()
    executor = _state.get("executor")
    return await loop.run_in_executor(executor, lambda: func(*args, **kwargs))


def _coerce_submission_shape(obj: Any) -> Dict[str, Any]:
    # If the payload is a bare list, assume it's the answers array
    if isinstance(obj, list):
        return {"answers": obj}
    if not isinstance(obj, dict):
        raise HTTPException(status_code=400, detail="Invalid submission payload; expected JSON object or array")
    # Accept common alternative keys
    if "answers" not in obj:
        if "items" in obj and isinstance(obj["items"], list):
            obj = {**obj, "answers": obj["items"]}
        elif "data" in obj and isinstance(obj["data"], list):
            obj = {**obj, "answers": obj["data"]}
    answers = obj.get("answers")
    if not isinstance(answers, list) or len(answers) == 0:
        raise HTTPException(status_code=400, detail="No answers provided; expected non-empty 'answers' array")
    return obj


def _write_team_xlsx(results_dir: str, participant_id: str, questions: List[Dict[str, Any]]) -> None:
    logger.debug("Preparing XLSX for participant=%s in dir=%s", participant_id, results_dir)
    try:
        from openpyxl import Workbook
    except Exception as exc:
        logger.warning("openpyxl not available, skipping XLSX for %s: %s", participant_id, exc)
        return
    wb = Workbook()
    ws = wb.active
    ws.title = "Results"
    # Layout: for each question, reserve 8 rows (4 variants per model × 2 models)
    # First columns per question on the first row: Qid, submitted answer, correct answers, final score, inconsistent
    # Next 8 rows (one per variant): correctness, conciseness, completeness, score, comment
    from openpyxl.styles import PatternFill
    red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    yellow_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")

    # Write a header block legend at the top
    ws.append(["Qid", "submitted answer", "correct answers", "final score", "inconsistent", "suspicious",
               "variant model", "variant correctness", "variant conciseness", "variant completeness", "variant score", "variant comment"]) 
    for col in range(1, 13):
        try:
            ws.column_dimensions[chr(64 + col)].width = 24 if col in (2,3,12) else 18
        except Exception:
            pass

    # Track rows containing final scores for formula generation
    score_rows = []

    for q in questions:
        qid = q.get("question_id")
        submitted = q.get("submitted_answer", "")
        eval_data = q.get("evaluation", {})
        final_score = eval_data.get("score")
        inconsistent = bool(eval_data.get("inconsistent", False))
        # Build main prompt info: include the question text and expected answer
        expected_text = ""
        try:
            qinfo = _state.get("questions", {}).get(qid, {})
            qtext = qinfo.get("question", "")
            exp = qinfo.get("expected_answer", "")
            if qtext or exp:
                expected_text = f"Question: {qtext}\nExpected: {exp}"
        except Exception:
            expected_text = ""
        # Write the question summary row
        suspicious = bool(eval_data.get("needs_manual_review", False))
        ws.append([qid, submitted, expected_text, final_score, inconsistent, suspicious, None, None, None, None, None, None])
        if inconsistent:
            ws.cell(row=ws.max_row, column=1).fill = red_fill
        if suspicious:
            ws.cell(row=ws.max_row, column=1).fill = yellow_fill
        
        # Track the row number for formula (column D = final score)
        if final_score is not None:
            score_rows.append(ws.max_row)

        # Variants
        v_scores = eval_data.get("variant_scores", []) or []
        v_comments = eval_data.get("variant_comments", []) or []
        v_weighted = eval_data.get("variant_weighted", []) or []
        # Ensure 8 rows (4 variants per model × 2 models)
        max_rows = 8
        for i in range(max_rows):
            if i < len(v_scores):
                v = v_scores[i]
                comment = v_comments[i] if i < len(v_comments) else ""
                w = v_weighted[i] if i < len(v_weighted) else None
                model_name = v.get("model", "unknown")
                ws.append([None, None, None, None, None, None, model_name, v.get("correctness"), v.get("conciseness"), v.get("completeness"), w, comment])
            else:
                ws.append([None]*12)
    
    # Add summary section at the end with Excel formulas
    if score_rows:
        # Add empty row for separation
        ws.append([None]*12)
        
        # Add summary rows with bold styling
        from openpyxl.styles import Font
        bold_font = Font(bold=True, size=12)
        green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
        
        # Build Excel formula ranges for SUM and AVERAGE
        # Column D contains the final scores
        score_cells = [f"D{row}" for row in score_rows]
        sum_formula = f"=SUM({','.join(score_cells)})"
        avg_formula = f"=AVERAGE({','.join(score_cells)})"
        count_formula = f"=COUNTA({','.join(score_cells)})"
        
        # Total sum row with formula
        ws.append(["TOTAL SCORE", None, None, None, None, None, None, None, None, None, None, None])
        total_row = ws.max_row
        ws.cell(row=total_row, column=4).value = sum_formula
        ws.cell(row=total_row, column=1).font = bold_font
        ws.cell(row=total_row, column=1).fill = green_fill
        ws.cell(row=total_row, column=4).font = bold_font
        ws.cell(row=total_row, column=4).fill = green_fill
        
        # Average score row with formula
        ws.append(["AVERAGE SCORE", None, None, None, None, None, None, None, None, None, None, None])
        avg_row = ws.max_row
        ws.cell(row=avg_row, column=4).value = avg_formula
        ws.cell(row=avg_row, column=1).font = bold_font
        ws.cell(row=avg_row, column=1).fill = green_fill
        ws.cell(row=avg_row, column=4).font = bold_font
        ws.cell(row=avg_row, column=4).fill = green_fill
        
        # Number of questions with formula
        ws.append(["NUMBER OF QUESTIONS", None, None, None, None, None, None, None, None, None, None, None])
        count_row = ws.max_row
        ws.cell(row=count_row, column=4).value = count_formula
        ws.cell(row=count_row, column=1).font = bold_font
        ws.cell(row=count_row, column=4).font = bold_font
    
    # Column widths already set in header block above; no further header-based sizing here
    os.makedirs(results_dir, exist_ok=True)
    xlsx_path = os.path.abspath(os.path.join(results_dir, f"{participant_id}.xlsx"))
    try:
        wb.save(xlsx_path)
        logger.info("Wrote XLSX: %s", xlsx_path)
    except Exception as exc:
        logger.exception("Failed to write XLSX %s: %s", xlsx_path, exc)


@app.post("/grade")
async def grade_submission(
    request: Request,
    use_llm: bool = Query(True, description="Use OpenAI LLM instead of heuristics"),
    write_files: bool = Query(True, description="Write JSON and update summary.csv on disk"),
    x_submission_token: Optional[str] = Header(None, alias="X-Submission-Token"),
) -> Dict[str, Any]:
    # Check content length FIRST before any other processing
    content_length = request.headers.get("content-length")
    if content_length and int(content_length) > MAX_SUBMISSION_SIZE:
        raise HTTPException(
            status_code=413,
            detail=f"Submission too large. Maximum size: {MAX_SUBMISSION_SIZE / (1024*1024):.1f}MB"
        )

    token, info = _require_token_and_team(x_submission_token)
    team = info.get("team", "unknown")

    questions = _state.get("questions") or {}
    if not questions:
        raise HTTPException(status_code=500, detail="Questions not loaded")

    try:
        body = await request.json()
    except Exception:
        raise HTTPException(status_code=400, detail="Invalid JSON body")

    sub = _coerce_submission_shape(body)
    sub = dict(sub)
    sub["participant_id"] = team

    if use_llm and not os.getenv("OPENAI_API_KEY"):
        raise HTTPException(status_code=400, detail="Missing OPENAI_API_KEY on server. Set it or call with use_llm=false.")

    try:
        result = await _run_in_executor(
            evaluate_submission,
            questions,
            sub,
            use_llm,
            DEFAULT_MODEL,
            FIXED_WORKERS,
            None,
            int(os.getenv("SELF_CONSISTENCY_RUNS", "3")),
            True,  # dual_model=True
        )
    except HTTPException:
        raise
    except Exception as exc:
        logger.exception("Grading failed for team %s (single)", team)
        raise HTTPException(status_code=400, detail=str(exc))

    if write_files:
        try:
            await _run_in_executor(write_results, result, RESULTS_DIR)
            csv_path = os.path.join(RESULTS_DIR, "summary.csv")
            rows: List[Dict[str, Any]] = []
            pid = result.get("participant_id") or "unknown"
            for q in result.get("questions", []):
                eval_data = q["evaluation"]
                rows.append(
                    {
                        "participant_id": pid,
                        "question_id": q["question_id"],
                        "completeness": eval_data["completeness"],
                        "conciseness": eval_data["conciseness"],
                        "correctness": eval_data["correctness"],
                        "score": eval_data["score"],
                    }
                )
            # Write CSV using shared helper on executor
            await _run_in_executor(write_summary_csv, csv_path, rows)
            # Also write XLSX per participant
            await _run_in_executor(_write_team_xlsx, RESULTS_DIR, pid, result.get("questions", []))
            # Mark token as used and persist
            info["used"] = True
            _state["token_to_info"][token] = info
            _persist_tokens(_state["token_to_info"])
            # Send confirmation email AFTER all processing is complete (best-effort)
            try:
                _send_confirmation_email(info.get("email", ""), team, sub)
            except Exception:
                logger.exception("Email confirmation failed for team %s", team)
        except Exception as exc:
            logger.exception("Failed to write results for participant %s", pid)
            raise HTTPException(status_code=500, detail=f"Failed to write results: {exc}")

    return result


@app.post("/submit")
async def submit_answers(
    request: Request,
    x_team_token: Optional[str] = Header(None, alias="X-Team-Token"),
) -> Dict[str, Any]:
    """
    Accept submission and process asynchronously.
    Returns 202 Accepted immediately, then processes in background.
    """
    # Check content length FIRST before any other processing
    content_length = request.headers.get("content-length")
    if content_length and int(content_length) > MAX_SUBMISSION_SIZE:
        raise HTTPException(
            status_code=413,
            detail=f"Submission too large. Maximum size: {MAX_SUBMISSION_SIZE / (1024*1024):.1f}MB"
        )

    # Validate token and get team info
    token, info = _require_token_and_team(x_team_token)
    team = info.get("team", "unknown")
    email = info.get("email", "")

    # Load questions
    questions = _state.get("questions") or {}
    if not questions:
        raise HTTPException(status_code=500, detail="Questions not loaded")

    # Parse and validate submission format
    try:
        body = await request.json()
    except Exception:
        raise HTTPException(status_code=400, detail="Invalid JSON body")

    try:
        sub = _coerce_submission_shape(body)
    except HTTPException:
        raise
    
    # Quick validation of submission structure
    sub = dict(sub)
    sub["participant_id"] = team
    
    # Validate we have OpenAI API key if needed
    if not os.getenv("OPENAI_API_KEY"):
        raise HTTPException(status_code=400, detail="Missing OPENAI_API_KEY on server")

    # Mark token as used IMMEDIATELY (before background processing)
    # This prevents duplicate submissions while grading
    info["used"] = True
    _state["token_to_info"][token] = info
    _persist_tokens(_state["token_to_info"])
    
    logger.info("Submission accepted for team=%s, starting background grading", team)
    
    # Launch background processing
    asyncio.create_task(_process_submission_background(team, email, sub, questions, body))
    
    # Return 202 Accepted immediately
    return {
        "status": "accepted",
        "message": f"Submission received for team {team}. Grading in progress. You will receive an email when complete.",
        "participant_id": team,
        "answers_count": len(sub.get("answers", []))
    }


async def _process_submission_background(
    team: str,
    email: str, 
    submission: Dict[str, Any],
    questions: Dict[str, Any],
    original_body: Dict[str, Any]
) -> None:
    """
    Process submission in background: grade, write files, send email.
    This runs asynchronously after returning 202 to the client.
    """
    try:
        logger.info("Background grading started for team=%s", team)
        
        # Run grading (CPU intensive, so use executor)
        result = await _run_in_executor(
            evaluate_submission,
            questions,
            submission,
            True,  # use_llm
            DEFAULT_MODEL,
            FIXED_WORKERS,
            None,
            int(os.getenv("SELF_CONSISTENCY_RUNS", "3")),
            True,  # dual_model=True
        )
        
        logger.info("Background grading completed for team=%s", team)
        
        # Write results to disk
        try:
            await _run_in_executor(write_results, result, RESULTS_DIR)
            
            # Write CSV summary
            csv_path = os.path.join(RESULTS_DIR, "summary.csv")
            rows: List[Dict[str, Any]] = []
            pid = result.get("participant_id") or "unknown"
            for q in result.get("questions", []):
                eval_data = q["evaluation"]
                rows.append({
                    "participant_id": pid,
                    "question_id": q["question_id"],
                    "completeness": eval_data["completeness"],
                    "conciseness": eval_data["conciseness"],
                    "correctness": eval_data["correctness"],
                    "score": eval_data["score"],
                })
            await _run_in_executor(write_summary_csv, csv_path, rows)
            
            # Write XLSX per participant
            await _run_in_executor(_write_team_xlsx, RESULTS_DIR, pid, result.get("questions", []))
            
            logger.info("Results written for team=%s", team)
            
        except Exception as exc:
            logger.exception("Failed to write results for team=%s", team)
        
        # Send confirmation email (best-effort)
        try:
            if email:
                _send_confirmation_email(email, team, original_body)
                logger.info("Confirmation email sent to team=%s (%s)", team, email)
            else:
                logger.warning("No email address for team=%s, skipping confirmation", team)
        except Exception as exc:
            logger.exception("Failed to send confirmation email for team=%s", team)
        
        logger.info("Background processing completed successfully for team=%s", team)
        
    except Exception as exc:
        logger.exception("Background processing failed for team=%s", team)
        # Send error email if possible
        try:
            if email:
                _send_error_email(email, team, str(exc))
        except Exception:
            logger.exception("Failed to send error email for team=%s", team)


def _send_error_email(to_addr: str, participant_id: str, error_message: str) -> None:
    """Send an email notification when grading fails."""
    smtp_host = os.getenv("SMTP_HOST")
    smtp_from = os.getenv("SMTP_FROM")
    if not smtp_host or not smtp_from:
        logger.debug("Email disabled: SMTP_HOST/SMTP_FROM not configured")
        return
    
    smtp_port = int(os.getenv("SMTP_PORT", "587"))
    smtp_user = os.getenv("SMTP_USER", "")
    smtp_password = os.getenv("SMTP_PASSWORD", "")
    use_ssl = (smtp_port == 465)
    
    msg = EmailMessage()
    msg["Subject"] = f"Submission Error - {participant_id}"
    msg["From"] = smtp_from
    msg["To"] = to_addr
    
    text_body = f"""Hello {participant_id},

Unfortunately, there was an error processing your submission.

Error: {error_message}

Please contact the competition organizers for assistance.

Best regards,
The Argusa Data Challenge Team
"""
    
    msg.set_content(text_body)
    
    try:
        if use_ssl:
            context = ssl.create_default_context()
            with smtplib.SMTP_SSL(smtp_host, smtp_port, context=context, timeout=30) as s:
                if smtp_user and smtp_password:
                    s.login(smtp_user, smtp_password)
                s.send_message(msg)
        else:
            with smtplib.SMTP(smtp_host, smtp_port, timeout=30) as s:
                s.ehlo()
                try:
                    s.starttls(context=ssl.create_default_context())
                    s.ehlo()
                except Exception:
                    pass
                if smtp_user and smtp_password:
                    s.login(smtp_user, smtp_password)
                s.send_message(msg)
        logger.info("Sent error email to %s", to_addr)
    except Exception as exc:
        logger.exception("Failed to send error email to %s: %s", to_addr, exc)


@app.post("/grade-batch")
async def grade_batch(
    request: Request,
    use_llm: bool = Query(True),
    write_files: bool = Query(True),
    x_submission_token: Optional[str] = Header(None, alias="X-Submission-Token"),
) -> Dict[str, Any]:
    # Check content length FIRST before any other processing
    content_length = request.headers.get("content-length")
    if content_length and int(content_length) > MAX_SUBMISSION_SIZE:
        raise HTTPException(
            status_code=413,
            detail=f"Submission too large. Maximum size: {MAX_SUBMISSION_SIZE / (1024*1024):.1f}MB"
        )

    token, info = _require_token_and_team(x_submission_token)
    team = info.get("team", "unknown")

    questions = _state.get("questions") or {}
    if not questions:
        raise HTTPException(status_code=500, detail="Questions not loaded")

    try:
        items = await request.json()
    except Exception:
        raise HTTPException(status_code=400, detail="Invalid JSON body")
    if not isinstance(items, list):
        raise HTTPException(status_code=400, detail="Batch body must be a JSON array of submissions")

    results: List[Dict[str, Any]] = []

    if write_files:
        _ensure_results_dir()
        csv_path = os.path.join(RESULTS_DIR, "summary.csv")
        all_rows: List[Dict[str, Any]] = []

    for item in items:
        try:
            sub = _coerce_submission_shape(item)
        except HTTPException as he:
            logger.error("Invalid submission shape: %s", he.detail)
            results.append({"error": he.detail})
            continue
        sub = dict(sub)
        sub["participant_id"] = team
        if use_llm and not os.getenv("OPENAI_API_KEY"):
            results.append({"participant_id": sub.get("participant_id", "unknown"), "error": "Missing OPENAI_API_KEY on server. Set it or call with use_llm=false."})
            continue
        try:
            result = await _run_in_executor(
                evaluate_submission,
                questions,
                sub,
                use_llm,
                DEFAULT_MODEL,
                FIXED_WORKERS,
                None,
                int(os.getenv("SELF_CONSISTENCY_RUNS", "3")),
                True,  # dual_model=True
            )
            results.append(result)
        except Exception as exc:
            logger.exception("Grading failed for team %s (batch)", team)
            results.append({"participant_id": sub.get("participant_id", "unknown"), "error": str(exc)})
            continue

        if write_files:
            try:
                await _run_in_executor(write_results, result, RESULTS_DIR)
                pid = result.get("participant_id") or "unknown"
                for q in result.get("questions", []):
                    eval_data = q["evaluation"]
                    all_rows.append(
                        {
                            "participant_id": pid,
                            "question_id": q["question_id"],
                            "completeness": eval_data["completeness"],
                            "conciseness": eval_data["conciseness"],
                            "correctness": eval_data["correctness"],
                            "score": eval_data["score"],
                        }
                    )
                # Also write XLSX per participant
                await _run_in_executor(_write_team_xlsx, RESULTS_DIR, pid, result.get("questions", []))
            except Exception as exc:
                logger.exception("Failed to write results for participant %s (batch)", pid)
                results.append({"participant_id": pid, "error": f"Failed to write results: {exc}"})

    if write_files:
        try:
            await _run_in_executor(write_summary_csv, csv_path, all_rows)
            # After batch, mark token used and persist
            info["used"] = True
            _state["token_to_info"][token] = info
            _persist_tokens(_state["token_to_info"])
            # Send confirmation email AFTER all processing is complete (best-effort)
            try:
                _send_confirmation_email(info.get("email", ""), team, {"submissions": items})
            except Exception:
                logger.exception("Email confirmation failed for team %s (batch)", team)
        except Exception as exc:
            logger.exception("Failed to write summary CSV")
            raise HTTPException(status_code=500, detail=f"Failed to write summary: {exc}")

    return {"results": results}
